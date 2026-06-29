"""
Read / write Job_Tracker.xlsx - one tab per portal, multi-profile matching, apply queues.
"""

from __future__ import annotations

import re
from typing import Any, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

import config
from job_utils import JobRecord, dedup_key, detect_experience_level, format_date, format_dt, now_aest, parse_added_on
from profile_loader import load_profiles, profile_summary_rows
from resume_matcher import match_label, score_all_profiles, score_job_for_profile


class ExcelManager:
    """Manages Job_Tracker.xlsx with per-portal sheets and per-doctor profile tracking."""

    def __init__(self, path: str = config.EXCEL_FILE_PATH):
        self.path = path
        self.profiles: list[dict[str, Any]] = load_profiles()
        self.COL_MAP = {name: idx + 1 for idx, name in enumerate(config.SHEET_COLUMNS)}
        self.wb = self._load_or_create()
        self._index: dict[str, dict[str, tuple[int, str]]] = {}
        self._new_row_keys: dict[str, set[str]] = {
            s: set() for s in config.ALL_JOB_SOURCE_SHEETS
        }
        self._build_index()
        self.refresh_profiles_sheet()

    def _load_or_create(self):
        try:
            wb = load_workbook(self.path)
            self._migrate_sheets(wb)
            self._migrate_columns(wb)
            return wb
        except FileNotFoundError:
            wb = Workbook()
            default = wb.active
            wb.remove(default)
            for sheet in config.PLATFORM_SHEETS:
                ws = wb.create_sheet(sheet)
                ws.append(config.SHEET_COLUMNS)
            for name, cols in [
                (config.DAILY_SUMMARY_SHEET, config.SUMMARY_COLUMNS),
                (config.PROFILES_SHEET, config.PROFILES_COLUMNS),
                (config.ALL_JOBS_SHEET, config.ALL_JOBS_COLUMNS),
                (config.PROFILE_MATCHES_SHEET, self._profile_matches_header()),
                (config.APPLY_QUEUE_SHEET, config.APPLY_QUEUE_COLUMNS),
            ]:
                ws = wb.create_sheet(name)
                ws.append(cols)
            wb.save(self.path)
            return wb

    def _profile_matches_header(self) -> list[str]:
        cols = list(config.PROFILE_MATCHES_COLUMNS)
        for p in self.profiles:
            short = (p.get("name") or p.get("id", ""))[:25]
            cols.append(f"{short} %")
        return cols

    def _migrate_sheets(self, wb: Workbook) -> None:
        changed = False
        for sheet in config.PLATFORM_SHEETS:
            if sheet not in wb.sheetnames:
                wb.create_sheet(sheet).append(config.SHEET_COLUMNS)
                changed = True
        extra = [
            (config.DAILY_SUMMARY_SHEET, config.SUMMARY_COLUMNS),
            (config.PROFILES_SHEET, config.PROFILES_COLUMNS),
            (config.ALL_JOBS_SHEET, config.ALL_JOBS_COLUMNS),
            (config.PROFILE_MATCHES_SHEET, self._profile_matches_header()),
            (config.APPLY_QUEUE_SHEET, config.APPLY_QUEUE_COLUMNS),
        ]
        for name, cols in extra:
            if name not in wb.sheetnames:
                wb.create_sheet(name).append(cols)
                changed = True
        if changed:
            wb.save(self.path)

    def _migrate_columns(self, wb: Workbook) -> None:
        changed = False
        for sheet in config.ALL_JOB_SOURCE_SHEETS + [config.ALL_JOBS_SHEET]:
            if sheet not in wb.sheetnames:
                continue
            ws = wb[sheet]
            headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
            for col in ("Portal", "Best Profile"):
                if col not in headers:
                    insert_at = headers.index("Match %") + 1 if "Match %" in headers else ws.max_column + 1
                    ws.insert_cols(insert_at)
                    ws.cell(1, insert_at, col)
                    changed = True
                    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
            if "Experience Level" not in headers:
                insert_at = headers.index("Specialty") + 2 if "Specialty" in headers else 3
                ws.insert_cols(insert_at)
                ws.cell(1, insert_at, "Experience Level")
                changed = True
                headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
                title_col = headers.index("Job Title") + 1 if "Job Title" in headers else 1
                spec_col = headers.index("Specialty") + 1 if "Specialty" in headers else 2
                for row in range(2, ws.max_row + 1):
                    title = ws.cell(row, title_col).value or ""
                    specialty = ws.cell(row, spec_col).value or ""
                    level = detect_experience_level(f"{title} {specialty}")
                    if level:
                        ws.cell(row, insert_at, level)
        if changed:
            wb.save(self.path)

    def refresh_profiles_sheet(self) -> None:
        ws = self._ensure_custom_sheet(config.PROFILES_SHEET, config.PROFILES_COLUMNS)
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        for row in profile_summary_rows(self.profiles):
            ws.append(row)

    def _ensure_sheet(self, name: str) -> Worksheet:
        if name not in self.wb.sheetnames:
            ws = self.wb.create_sheet(name)
            ws.append(config.SHEET_COLUMNS)
        ws = self.wb[name]
        if ws.max_row == 0 or ws.cell(1, 1).value != config.SHEET_COLUMNS[0]:
            ws.delete_rows(1, ws.max_row)
            ws.append(config.SHEET_COLUMNS)
        return ws

    def _ensure_custom_sheet(self, name: str, columns: list[str]) -> Worksheet:
        if name not in self.wb.sheetnames:
            ws = self.wb.create_sheet(name)
            ws.append(columns)
            return ws
        ws = self.wb[name]
        if ws.max_row == 0 or ws.cell(1, 1).value != columns[0]:
            ws.delete_rows(1, ws.max_row)
            ws.append(columns)
        return ws

    def _ensure_summary_sheet(self) -> Worksheet:
        return self._ensure_custom_sheet(config.DAILY_SUMMARY_SHEET, config.SUMMARY_COLUMNS)

    def _build_index(self) -> None:
        self._index = {}
        for sheet in config.ALL_JOB_SOURCE_SHEETS:
            if sheet not in self.wb.sheetnames:
                continue
            ws = self.wb[sheet]
            self._index[sheet] = {}
            for row in range(2, ws.max_row + 1):
                title = safe_cell(ws, row, "Job Title", self.COL_MAP)
                hospital = safe_cell(ws, row, "Hospital", self.COL_MAP)
                link = safe_cell(ws, row, "Apply Link", self.COL_MAP)
                if title or hospital:
                    self._index[sheet][dedup_key(hospital, title)] = (row, link)

    def sheet_job_count(self, sheet: str) -> int:
        if sheet not in self.wb.sheetnames:
            return 0
        return max(0, self.wb[sheet].max_row - 1)

    def _score_job(self, job: JobRecord, portal_sheet: str) -> tuple[str, str, int]:
        if not self.profiles:
            return ("", "", job.match_pct or 0)
        scored = score_all_profiles(
            self.profiles,
            job.title,
            job.description,
            job.specialty,
            job.location,
            job.state,
        )
        best = scored[0]
        return (best["profile_id"], best["profile_name"], best["match_pct"])

    def add_jobs(self, sheet: str, jobs: list[JobRecord]) -> int:
        ws = self._ensure_sheet(sheet)
        if sheet not in self._index:
            self._index[sheet] = {}
        if sheet not in self._new_row_keys:
            self._new_row_keys[sheet] = set()

        added = 0
        added_at = format_dt(now_aest())

        for job in sorted(jobs, key=lambda j: j.match_pct, reverse=True):
            key = dedup_key(job.hospital, job.title)
            posted_str = format_date(job.posted_date) if job.posted_date else format_date(now_aest())

            profile_id, profile_name, match_pct = self._score_job(job, sheet)
            if job.match_pct and job.match_pct > match_pct:
                match_pct = job.match_pct

            label = match_label(match_pct)
            notes = job.notes or ""
            if label == "High Match" and "High Match" not in notes:
                notes = f"High Match | {notes}".strip(" |")

            if key in self._index[sheet]:
                row_num, old_link = self._index[sheet][key]
                if job.apply_link and job.apply_link != old_link:
                    ws.cell(row_num, self.COL_MAP["Apply Link"], job.apply_link)
                ws.cell(row_num, self.COL_MAP["Best Profile"], profile_name)
                ws.cell(row_num, self.COL_MAP["Match %"], match_pct)
                ws.cell(row_num, self.COL_MAP["Portal"], sheet)
                self._index[sheet][key] = (row_num, job.apply_link)
                continue

            row_data = [
                job.title,
                job.specialty,
                job.experience_level,
                job.hospital,
                job.location,
                job.state,
                job.salary,
                posted_str,
                added_at,
                job.apply_link,
                sheet,
                profile_name,
                match_pct,
                config.STATUS_NEW,
                "",
                notes,
            ]
            ws.append(row_data)
            self._index[sheet][key] = (ws.max_row, job.apply_link)
            self._new_row_keys[sheet].add(key)
            added += 1

        return added

    def rebuild_all_jobs_sheet(self) -> int:
        ws = self._ensure_custom_sheet(config.ALL_JOBS_SHEET, config.ALL_JOBS_COLUMNS)
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        count = 0
        for sheet in config.ALL_JOB_SOURCE_SHEETS:
            if sheet not in self.wb.sheetnames:
                continue
            src = self.wb[sheet]
            for row in range(2, src.max_row + 1):
                title = src.cell(row, 1).value
                if not title:
                    continue
                ws.append([src.cell(row, c).value for c in range(1, len(config.SHEET_COLUMNS) + 1)])
                count += 1
        return count

    def rebuild_profile_matches_sheet(self) -> int:
        header = self._profile_matches_header()
        ws = self._ensure_custom_sheet(config.PROFILE_MATCHES_SHEET, header)
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        count = 0
        for sheet in config.ALL_JOB_SOURCE_SHEETS:
            if sheet not in self.wb.sheetnames:
                continue
            src = self.wb[sheet]
            for row in range(2, src.max_row + 1):
                title = src.cell(row, 1).value
                if not title:
                    continue
                hospital = safe_cell(src, row, "Hospital", self.COL_MAP)
                state = safe_cell(src, row, "State", self.COL_MAP)
                link = safe_cell(src, row, "Apply Link", self.COL_MAP)
                best_name = safe_cell(src, row, "Best Profile", self.COL_MAP)
                best_pct = src.cell(row, self.COL_MAP["Match %"]).value or 0
                row_data = [title, sheet, hospital, state, link, best_name, best_pct]
                desc = (
                    f"{title} {safe_cell(src, row, 'Specialty', self.COL_MAP)} "
                    f"{safe_cell(src, row, 'Experience Level', self.COL_MAP)} {hospital}"
                )
                for p in self.profiles:
                    pct = score_job_for_profile(
                        p, str(title), desc,
                        str(safe_cell(src, row, "Specialty", self.COL_MAP)),
                        str(safe_cell(src, row, "Location", self.COL_MAP)),
                        str(state),
                    )
                    row_data.append(pct)
                ws.append(row_data)
                count += 1
        return count

    def update_sheet_statuses(self, sheet: Optional[str] = None) -> None:
        sheets = [sheet] if sheet else config.ALL_JOB_SOURCE_SHEETS
        today = now_aest().date()
        for sheet_name in sheets:
            if sheet_name not in self.wb.sheetnames:
                continue
            ws = self.wb[sheet_name]
            new_keys = self._new_row_keys.get(sheet_name, set())
            for row in range(2, ws.max_row + 1):
                title = safe_cell(ws, row, "Job Title", self.COL_MAP)
                hospital = safe_cell(ws, row, "Hospital", self.COL_MAP)
                if not title and not hospital:
                    continue
                key = dedup_key(hospital, title)
                applied = safe_cell(ws, row, "Applied?", self.COL_MAP).strip().lower()
                if applied in ("y", "yes", "applied"):
                    ws.cell(row, self.COL_MAP["Status"], config.STATUS_APPLIED)
                    continue
                if key in new_keys:
                    ws.cell(row, self.COL_MAP["Status"], config.STATUS_NEW)
                    continue
                added_dt = parse_added_on(safe_cell(ws, row, "Job Added On", self.COL_MAP))
                if not added_dt:
                    continue
                days = (today - added_dt.date()).days
                if days <= 0:
                    status = config.STATUS_NEW
                elif days == 1:
                    status = "1 day old"
                elif days == 2:
                    status = "2 days old"
                else:
                    status = f"{days} days old"
                ws.cell(row, self.COL_MAP["Status"], status)

    def append_summary(self, platform: str, method: str, new_jobs: int, total_jobs: int, error: str = "") -> None:
        ws = self._ensure_summary_sheet()
        run = now_aest()
        ws.append([
            run.strftime("%d-%m-%Y"),
            run.strftime("%H:%M"),
            platform,
            method,
            new_jobs,
            total_jobs,
            error or "No",
        ])

    def rebuild_apply_queue(self) -> int:
        ws_queue = self._ensure_custom_sheet(config.APPLY_QUEUE_SHEET, config.APPLY_QUEUE_COLUMNS)
        if ws_queue.max_row > 1:
            ws_queue.delete_rows(2, ws_queue.max_row - 1)

        candidates: list[dict] = []
        for sheet in config.ALL_JOB_SOURCE_SHEETS:
            if sheet not in self.wb.sheetnames:
                continue
            ws = self.wb[sheet]
            for row in range(2, ws.max_row + 1):
                title = ws.cell(row, 1).value
                if not title:
                    continue
                applied = safe_cell(ws, row, "Applied?", self.COL_MAP).strip().lower()
                status = safe_cell(ws, row, "Status", self.COL_MAP)
                if applied in ("y", "yes", "applied") or status == config.STATUS_EXPIRED:
                    continue
                try:
                    match_pct = int(float(ws.cell(row, self.COL_MAP["Match %"]).value or 0))
                except (ValueError, TypeError):
                    match_pct = 0
                candidates.append({
                    "profile": safe_cell(ws, row, "Best Profile", self.COL_MAP),
                    "sheet": sheet,
                    "title": str(title),
                    "specialty": safe_cell(ws, row, "Specialty", self.COL_MAP),
                    "experience": safe_cell(ws, row, "Experience Level", self.COL_MAP),
                    "hospital": safe_cell(ws, row, "Hospital", self.COL_MAP),
                    "location": safe_cell(ws, row, "Location", self.COL_MAP),
                    "state": safe_cell(ws, row, "State", self.COL_MAP),
                    "link": safe_cell(ws, row, "Apply Link", self.COL_MAP),
                    "match_pct": match_pct,
                    "status": status,
                    "notes": safe_cell(ws, row, "Notes", self.COL_MAP),
                })

        candidates.sort(key=lambda c: c["match_pct"], reverse=True)
        for rank, c in enumerate(candidates[: config.APPLY_QUEUE_SIZE], start=1):
            ws_queue.append([
                rank,
                c["profile"],
                c["sheet"],
                c["title"],
                c["specialty"],
                c["experience"],
                c["hospital"],
                c["location"],
                c["state"],
                c["link"],
                c["match_pct"],
                match_label(c["match_pct"]),
                c["status"],
                "",
                c["notes"],
            ])

        self._rebuild_per_profile_queues(candidates)
        return min(len(candidates), config.APPLY_QUEUE_SIZE)

    def _rebuild_per_profile_queues(self, candidates: list[dict]) -> None:
        for profile in self.profiles:
            pid = profile.get("id", "profile")
            pname = profile.get("name", pid)
            safe_name = re.sub(r"[^\w]", "_", pid)[:28]
            sheet_name = f"Queue_{safe_name}"[:31]
            cols = config.APPLY_QUEUE_COLUMNS
            ws = self._ensure_custom_sheet(sheet_name, cols)
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row - 1)
            profile_jobs = [c for c in candidates if c["profile"] == pname]
            profile_jobs.sort(key=lambda c: c["match_pct"], reverse=True)
            for rank, c in enumerate(profile_jobs[:15], start=1):
                ws.append([
                    rank, pname, c["sheet"], c["title"], c["specialty"], c["experience"],
                    c["hospital"], c["location"], c["state"], c["link"],
                    c["match_pct"], match_label(c["match_pct"]), c["status"], "", c["notes"],
                ])

    def get_follow_up_jobs(self) -> list[dict]:
        follow_ups: list[dict] = []
        cutoff = now_aest().date()
        for sheet in config.ALL_JOB_SOURCE_SHEETS:
            if sheet not in self.wb.sheetnames:
                continue
            ws = self.wb[sheet]
            for row in range(2, ws.max_row + 1):
                applied = safe_cell(ws, row, "Applied?", self.COL_MAP).strip().lower()
                if applied not in ("y", "yes", "applied"):
                    continue
                notes = safe_cell(ws, row, "Notes", self.COL_MAP).lower()
                if any(x in notes for x in ("response", "interview", "rejected")):
                    continue
                added_dt = parse_added_on(safe_cell(ws, row, "Job Added On", self.COL_MAP))
                if not added_dt:
                    continue
                days = (cutoff - added_dt.date()).days
                if days >= config.FOLLOW_UP_AFTER_DAYS:
                    follow_ups.append({
                        "sheet": sheet,
                        "title": safe_cell(ws, row, "Job Title", self.COL_MAP),
                        "hospital": safe_cell(ws, row, "Hospital", self.COL_MAP),
                        "profile": safe_cell(ws, row, "Best Profile", self.COL_MAP),
                        "days": days,
                        "link": safe_cell(ws, row, "Apply Link", self.COL_MAP),
                    })
        return follow_ups

    def auto_column_width(self) -> None:
        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
            for col_idx in range(1, ws.max_column + 1):
                max_len = 10
                col_letter = get_column_letter(col_idx)
                for row in range(1, min(ws.max_row + 1, 150)):
                    val = ws.cell(row, col_idx).value
                    if val is not None:
                        max_len = max(max_len, len(str(val)))
                ws.column_dimensions[col_letter].width = min(max_len + 2, 55)

    def save(self) -> None:
        self.auto_column_width()
        self.wb.save(self.path)




def safe_cell(ws: Worksheet, row: int, col_name: str, col_map: dict[str, int] | None = None) -> str:
    cmap = col_map or {name: idx + 1 for idx, name in enumerate(config.SHEET_COLUMNS)}
    if col_name not in cmap:
        return ""
    val = ws.cell(row, cmap[col_name]).value
    return "" if val is None else str(val).strip()
