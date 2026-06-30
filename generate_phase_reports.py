#!/usr/bin/env python3
"""Generate Phase 2 validation report and Phase 4 extraction strategy report."""

from __future__ import annotations

import json
from collections import defaultdict
from pathlib import Path

import openpyxl

import config
from extraction_runner import get_method_rankings
from job_validator import validate_job_against_profiles
from profile_loader import load_profiles


def _load_workbook_jobs() -> list[dict]:
    wb = openpyxl.load_workbook(config.EXCEL_FILE_PATH, read_only=True, data_only=True)
    jobs: list[dict] = []
    col = {name: idx for idx, name in enumerate(config.SHEET_COLUMNS)}

    def cell(row, name, default=""):
        idx = col.get(name)
        if idx is None or idx >= len(row):
            return default
        v = row[idx]
        return default if v is None else v

    for sheet in config.ALL_JOB_SOURCE_SHEETS:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            jobs.append({
                "sheet": sheet,
                "title": str(cell(row, "Job Title")),
                "specialty": str(cell(row, "Specialty")),
                "experience": str(cell(row, "Experience Level")),
                "hospital": str(cell(row, "Hospital")),
                "location": str(cell(row, "Location")),
                "state": str(cell(row, "State")),
                "best_profile": str(cell(row, "Best Profile")),
                "match_pct": cell(row, "Match %", 0),
                "extract_method": str(cell(row, "Extraction Method Used")),
            })
    wb.close()
    return jobs


def generate_phase2_report(jobs: list[dict], profiles: list) -> str:
    by_platform: dict[str, list] = defaultdict(list)
    flagged = 0
    fuzzy = 0
    removed = 0
    reviewed = len(jobs)

    for job in jobs:
        result = validate_job_against_profiles(
            profiles,
            job["title"],
            "",
            job["specialty"],
            job["location"],
            job["state"],
            job["experience"],
        )
        job["validation"] = result
        by_platform[job["sheet"]].append(result)
        if result.flags:
            flagged += 1
        if result.title_match_type == "fuzzy":
            fuzzy += 1
        if not result.passed:
            removed += 1

    lines = [
        "# Phase 2: Data Validation Report",
        "",
        f"**Records reviewed:** {reviewed}",
        f"**Flagged (validation warnings):** {flagged}",
        f"**Fuzzy title matches:** {fuzzy}",
        f"**Would be filtered** (below threshold {config.VALIDATION_CONFIDENCE_THRESHOLD}%): {removed}",
        f"**Filter active:** {config.VALIDATION_FILTER_BELOW_THRESHOLD}",
        "",
        "## Validation Logic Implemented",
        "",
        "Module: `job_validator.py` — runs after fetch, before Excel write.",
        "",
        "| Component | Weight | Notes |",
        "|-----------|--------|-------|",
        "| Title/role relevance | 40% | Fuzzy match of job title to profile specialty + level (proxy for identity; job boards have no doctor names) |",
        "| Specialty alignment | 25% | Detected job specialty vs profile specialty |",
        "| Location alignment | 20% | Job state vs profile preferred_states |",
        "| Experience level | 15% | Job level vs profile experience_level |",
        "| License/registration | N/A | Not exposed on AU hospital job portals — weight redistributed |",
        "",
        f"**Confidence threshold:** {config.VALIDATION_CONFIDENCE_THRESHOLD}% (configurable in `config.py`)",
        "",
        "## Per-Platform Summary",
        "",
        "| Platform | Reviewed | Flagged | Fuzzy | Below threshold |",
        "|----------|----------|---------|-------|-----------------|",
    ]

    for sheet in config.ALL_JOB_SOURCE_SHEETS:
        results = by_platform.get(sheet, [])
        if not results:
            lines.append(f"| {sheet} | 0 | — | — | — |")
            continue
        f = sum(1 for r in results if r.flags)
        fz = sum(1 for r in results if r.title_match_type == "fuzzy")
        low = sum(1 for r in results if not r.passed)
        lines.append(f"| {sheet} | {len(results)} | {f} | {fz} | {low} |")

    lines.extend([
        "",
        "## False-Positive Patterns",
        "",
        "- **Common specialty overlap:** General Medicine jobs matched to multiple profiles",
        "- **Fuzzy title matches:** Short titles (e.g. 'Registrar') match multiple specialties",
        "- **Location gaps:** Jobs with empty state field score low on location despite being AU hospital jobs",
        "",
        "## Sample Flagged Records",
        "",
    ])

    count = 0
    for job in jobs:
        r = job["validation"]
        if r.flags and count < 15:
            lines.append(f"- **{job['sheet']}** | {job['title'][:60]} | Profile: {r.profile_name} | Match: {r.match_pct}% | Flags: {', '.join(r.flags)}")
            count += 1

    return "\n".join(lines) + "\n"


def generate_phase4_report() -> str:
    rankings = get_method_rankings()
    health_path = config.LOGS_DIR / "portal_health.json"
    health = {}
    if health_path.exists():
        health = json.loads(health_path.read_text(encoding="utf-8"))

    lines = [
        "# Platform Extraction Strategy Report",
        "",
        "Generated from `extraction_runner.py`, `logs/extraction_stats.json`, and Phase 1 diagnosis.",
        "",
        "## Method Ranking Table",
        "",
        "| Platform | Method 1 | Method 2 | Working? | Best Method | Worst Method | Reason Best | Reason Worst |",
        "|----------|----------|----------|----------|-------------|--------------|-------------|--------------|",
    ]

    method_matrix = {
        "smartjobs_qld": ("playwright", "static/API", "N", "playwright", "static", "Only configured method; needs API/WAF bypass", "Static blocked by WAF"),
        "jobs_nt": ("playwright", "static", "N", "playwright", "static", "SPA needs browser", "Returns no results for current query"),
        "careers_vic": ("playwright", "static", "N", "playwright", "static", "JS portal", "404 URL — both fail"),
        "monash_health": ("playwright", "static", "Y", "playwright", "static", "23 jobs via Playwright", "Static returns empty shell"),
        "western_health": ("static", "playwright", "Y", "static", "playwright", "Static HTML sufficient — 12 jobs", "Playwright unnecessary"),
        "wa_health": ("playwright", "static", "N", "playwright", "static", "JS required", "Wrong URL — page not found"),
        "mercy_workday": ("workday_api", "mercury_html", "N", "mercury_html", "workday_api", "Mercy uses Mercury not Workday", "404 on configured Workday tenant"),
        "peninsula_health": ("static", "playwright", "N*", "playwright", "static", "Playwright finds 8+ jobs (not yet default)", "Static returns JS shell"),
        "the_womens": ("static", "successfactors", "N", "successfactors", "static", "Real jobs on SuccessFactors", "Timeout + wrong static URL"),
        "grampians_health": ("static", "playwright", "Y", "static", "playwright", "Static works — 14 jobs", "—"),
        "eastern_health": ("static", "playwright", "Y", "static", "playwright", "Static works — 12 jobs", "—"),
        "rch": ("static", "playwright", "Partial", "static", "playwright", "Static worked historically (7 stale rows)", "Latest run: connect timeout"),
        "ranzcog": ("playwright", "static", "N", "playwright", "static", "When not 403, page has articles", "403 Forbidden intermittent"),
        "racp": ("static", "playwright", "Y", "static", "playwright", "Static HTML — 41 jobs", "—"),
        "jobradars": ("static", "playwright", "N", "playwright", "static", "May work with browser", "403 bot block on HTTP"),
        "pageup": ("pageup_html", "playwright", "N", "playwright", "pageup_html", "JS PageUp sites need browser", "Dead DNS + false-positive links"),
    }

    for portal_key, cfg in config.PORTAL_CONFIG.items():
        sheet = cfg.get("sheet", portal_key)
        row = method_matrix.get(portal_key, ("primary", "fallback", "?", "primary", "fallback", "—", "—"))
        rank = rankings.get(portal_key, [])
        best_note = rank[0]["method"] if rank else row[3]
        lines.append(
            f"| {sheet} | {row[0]} | {row[1]} | {row[2]} | {best_note} | {row[4]} | {row[5]} | {row[6]} |"
        )

    disabled = health.get("disabled", [])
    lines.extend([
        "",
        "## Per-Platform Narrative",
        "",
        "See `phase1_missing_data_diagnosis.md` for detailed root-cause analysis.",
        "",
        f"**Auto-disabled portals ({len(disabled)}):** {', '.join(disabled)}",
        "",
        "Multi-method extraction is implemented in `extraction_runner.py`. Each run tries the configured",
        "primary method, then a fallback (static ↔ Playwright). Results are logged to",
        "`logs/extraction_stats.json`. Excel columns **Extraction Method Used** and",
        "**Method Reliability Note** record which method succeeded per job.",
        "",
        "## Recommendations",
        "",
        "### Keep using",
        "- **Static HTML** for Western Health, Grampians, Eastern Health, RACP",
        "- **Playwright** for Monash Health",
        "- **Multi-method fallback** chain (now automatic)",
        "",
        "### Deprecate / replace",
        "- **Mercy Workday API** → build Mercury scraper",
        "- **JobRadars HTTP** → blocked; use direct hospital sources",
        "- **PageUp link regex** on SA Health → false positives only",
        "",
        "### Engineering investment",
        "1. Peninsula Health — switch primary to Playwright (**low effort, high yield**)",
        "2. Careers VIC — fix 404 URL",
        "3. SmartJobs QLD — WAF bypass / API discovery",
        "4. The Women's — SuccessFactors parser",
        "5. Reset `portal_health.json` disabled list after fixes",
        "",
    ])
    return "\n".join(lines) + "\n"


def main() -> None:
    profiles = load_profiles()
    jobs = _load_workbook_jobs()
    root = Path(config.REPO_ROOT)

    p2 = generate_phase2_report(jobs, profiles)
    (root / "phase2_data_validation_report.md").write_text(p2, encoding="utf-8")
    print(f"Wrote phase2_data_validation_report.md ({len(jobs)} jobs reviewed)")

    p4 = generate_phase4_report()
    (root / "platform_extraction_strategy_report.md").write_text(p4, encoding="utf-8")
    print("Wrote platform_extraction_strategy_report.md")


if __name__ == "__main__":
    main()
