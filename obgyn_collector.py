#!/usr/bin/env python3
"""
OBGYN Job Tracker — dedicated collector for Obstetrics & Gynaecology Registrar roles.

Creates/updates OBGYN_Job_Tracker.xlsx with the same columns as Job_Tracker.xlsx.

1. Seeds from existing Job_Tracker.xlsx O&G rows
2. Scrapes project portals + Seek / Jora with O&G keywords
3. De-duplicates by Job URL
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Iterable
from urllib.parse import quote_plus, urlparse, urlunparse, parse_qsl, urlencode

from openpyxl import Workbook, load_workbook

import config
from job_utils import JobRecord, RunLogger, now_aest, format_dt, format_date
from scrapers import ALL_SCRAPERS
from scrapers.base import (
    absolute_url,
    fetch_html,
    fetch_with_playwright,
    parse_job_cards,
    soup_from_html,
    text,
)

PROFILE_TAG = "Obstetrics and Gynaecology Registrar / OBGYN"
OUTPUT_PATH = config.REPO_ROOT / "OBGYN_Job_Tracker.xlsx"
COLUMNS = list(config.SHEET_COLUMNS)

OBGYN_KEYWORDS = [
    "OBGYN Registrar",
    "O&G Registrar",
    "Obstetrics Gynaecology Registrar",
    "Obstetrics and Gynaecology Registrar",
    "Obstetrics & Gynaecology Registrar",
]

# Broader match terms used when filtering titles/descriptions
OBGYN_MATCH_TERMS = (
    "obgyn",
    "o&g",
    "o & g",
    "obstetric",
    "gynaecol",
    "gynecol",
    "ranzcog",
    "maternity",
    "midwifery medical",
    "women's health",
    "womens health",
)

# Role-level terms (prefer registrar-family for this profile, but keep consultants/fellows)
OBGYN_ROLE_TERMS = (
    "registrar",
    "senior registrar",
    "fellow",
    "consultant",
    "staff specialist",
    "obstetrician",
    "gynaecologist",
    "gynecologist",
    "pho",
    "principal house officer",
    "rmo",
    "medical officer",
)


def normalize_url(url: str) -> str:
    """Canonical URL for de-duplication (strip fragments, trailing slash, common tracking params)."""
    if not url:
        return ""
    url = url.strip()
    try:
        parsed = urlparse(url)
    except Exception:  # noqa: BLE001
        return url.lower().rstrip("/")
    query = [
        (k, v)
        for k, v in parse_qsl(parsed.query, keep_blank_values=True)
        if k.lower() not in {"utm_source", "utm_medium", "utm_campaign", "utm_term", "utm_content", "ref"}
    ]
    cleaned = parsed._replace(
        scheme=parsed.scheme.lower(),
        netloc=parsed.netloc.lower(),
        path=parsed.path.rstrip("/") or "/",
        query=urlencode(query),
        fragment="",
    )
    return urlunparse(cleaned)


def is_obgyn_job(title: str, specialty: str = "", description: str = "") -> bool:
    blob = f"{title} {specialty} {description}".lower()
    blob = (
        blob.replace("&amp;", "&")
        .replace("&#38;", "&")
        .replace(" and ", " ")
        .replace("  ", " ")
    )
    if not any(term in blob for term in OBGYN_MATCH_TERMS):
        return False
    # Prefer clinical O&G roles; drop pure midwifery/nursing if no doctor terms
    if re.search(r"\b(midwife|midwifery|nurse|nursing|enrolled nurse)\b", blob) and not any(
        t in blob for t in OBGYN_ROLE_TERMS
    ):
        return False
    return True


def _cell(ws, row: int, col_name: str, headers: list) -> str:
    try:
        idx = headers.index(col_name) + 1
    except ValueError:
        return ""
    val = ws.cell(row, idx).value
    return "" if val is None else str(val).strip()


def _make_row(
    title: str,
    apply_link: str,
    portal: str = "",
    specialty: str = "Obstetrics & Gynaecology",
    experience_level: str = "",
    hospital: str = "",
    location: str = "",
    state: str = "",
    salary: str = "",
    posted_date: str = "",
    job_added_on: str = "",
    best_profile: str = PROFILE_TAG,
    match_pct: int = 0,
    status: str = config.STATUS_NEW,
    applied: str = "",
) -> dict[str, str]:
    if not job_added_on:
        job_added_on = format_dt(now_aest())
    return {
        "Job Title": title,
        "Specialty": specialty or "Obstetrics & Gynaecology",
        "Experience Level": experience_level,
        "Hospital": hospital,
        "Location": location,
        "State": state,
        "Salary": salary,
        "Posted Date": posted_date,
        "Job Added On": job_added_on,
        "Apply Link": apply_link,
        "Portal": portal,
        "Best Profile": best_profile,
        "Match %": str(match_pct) if match_pct else "0",
        "Status": status,
        "Applied?": applied,
    }


def seed_from_job_tracker(logger: RunLogger) -> list[dict[str, str]]:
    """Import O&G rows from Job_Tracker.xlsx with all columns."""
    path = Path(config.EXCEL_FILE_PATH)
    if not path.exists():
        logger.log(f"Seed skipped — {path} not found")
        return []

    wb = load_workbook(path, read_only=True, data_only=True)
    rows: list[dict[str, str]] = []
    seen: set[str] = set()

    portal_sheets = [s for s in wb.sheetnames if s in config.ALL_JOB_SOURCE_SHEETS]
    if not portal_sheets:
        portal_sheets = [config.ALL_JOBS_SHEET] if config.ALL_JOBS_SHEET in wb.sheetnames else []

    for sheet_name in portal_sheets:
        ws = wb[sheet_name]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        if "Job Title" not in headers or "Apply Link" not in headers:
            continue

        for row in range(2, ws.max_row + 1):
            title = _cell(ws, row, "Job Title", headers)
            link = _cell(ws, row, "Apply Link", headers)
            specialty = _cell(ws, row, "Specialty", headers)
            if not title or not link:
                continue
            if not is_obgyn_job(title, specialty):
                continue
            key = normalize_url(link)
            if not key or key in seen:
                continue
            seen.add(key)
            rows.append(_make_row(
                title=title,
                apply_link=link,
                portal=_cell(ws, row, "Portal", headers) or sheet_name,
                specialty=specialty,
                experience_level=_cell(ws, row, "Experience Level", headers),
                hospital=_cell(ws, row, "Hospital", headers),
                location=_cell(ws, row, "Location", headers),
                state=_cell(ws, row, "State", headers),
                salary=_cell(ws, row, "Salary", headers),
                posted_date=_cell(ws, row, "Posted Date", headers),
                job_added_on=_cell(ws, row, "Job Added On", headers),
                best_profile=_cell(ws, row, "Best Profile", headers) or PROFILE_TAG,
                match_pct=int(float(_cell(ws, row, "Match %", headers) or "0")),
                status=_cell(ws, row, "Status", headers),
                applied=_cell(ws, row, "Applied?", headers),
            ))

    wb.close()
    logger.log(f"Seed: {len(rows)} OBGYN jobs from Job_Tracker.xlsx")
    return rows


def _row_from_job(job: JobRecord, portal: str) -> dict[str, str]:
    return _make_row(
        title=job.title,
        apply_link=job.apply_link,
        portal=portal,
        specialty=job.specialty or "Obstetrics & Gynaecology",
        experience_level=job.experience_level,
        hospital=job.hospital,
        location=job.location,
        state=job.state,
        salary=job.salary,
        posted_date=format_date(job.posted_date) if job.posted_date else "",
        match_pct=job.match_pct,
    )


def scrape_existing_portals(logger: RunLogger) -> list[dict[str, str]]:
    """Run existing scrapers and keep OBGYN-matching jobs."""
    found: list[dict[str, str]] = []
    for portal_key, scrape_fn in ALL_SCRAPERS.items():
        if not config.PLATFORMS_TO_RUN.get(portal_key, False):
            continue
        sheet = config.PORTAL_CONFIG.get(portal_key, {}).get("sheet", portal_key)
        try:
            logger.log(f"Scraping {portal_key} for OBGYN...")
            jobs: list[JobRecord] = scrape_fn()
            kept = 0
            for job in jobs:
                if not is_obgyn_job(job.title, job.specialty, job.description):
                    continue
                if not job.apply_link:
                    continue
                found.append(_row_from_job(job, sheet))
                kept += 1
            logger.log(f"  {portal_key}: {kept} OBGYN of {len(jobs)} raw")
        except Exception as exc:  # noqa: BLE001
            logger.log(f"  {portal_key} failed: {exc}")
    return found


def scrape_targeted_og_searches(logger: RunLogger) -> list[dict[str, str]]:
    """Hit key portals with explicit O&G registrar search URLs."""
    found: list[dict[str, str]] = []
    searches = [
        {
            "label": "Peninsula_Health_OG",
            "portal_key": "peninsula_health",
            "url": "https://careers.peninsulahealth.org.au/search/?q=obstetrics+gynaecology+registrar",
            "method": "playwright",
        },
        {
            "label": "The_Womens_OG",
            "portal_key": "the_womens",
            "url": "https://careers.thewomens.org.au/search/?q=obstetrics+gynaecology",
            "method": "playwright",
        },
        {
            "label": "Monash_Health_OG",
            "portal_key": "monash_health",
            "url": "https://careers.monashhealth.org/search/?q=obstetrics+gynaecology+registrar",
            "method": "playwright",
        },
        {
            "label": "Eastern_Health_OG",
            "portal_key": "eastern_health",
            "url": "https://careers.easternhealth.org.au/search/?q=obstetrics+gynaecology",
            "method": "static",
        },
        {
            "label": "Grampians_Health_OG",
            "portal_key": "grampians_health",
            "url": "https://careers.grampianshealth.com/search/?q=obstetrics+gynaecology",
            "method": "static",
        },
        {
            "label": "WA_Health_OG",
            "portal_key": "wa_health",
            "url": "https://medcareerswa.health.wa.gov.au/jobs/search?q=obstetrics+gynaecology",
            "method": "playwright",
        },
        {
            "label": "NSW_Health_OG",
            "portal_key": "pageup",
            "url": "https://jobs.health.nsw.gov.au/jobs/search?q=obstetrics+gynaecology",
            "method": "static",
        },
        {
            "label": "QLD_Health_OG",
            "portal_key": "pageup",
            "url": "https://careers.health.qld.gov.au/search/?q=obstetrics+gynaecology",
            "method": "playwright",
        },
        {
            "label": "Careers_VIC_OG",
            "portal_key": "careers_vic",
            "url": "https://careers.vic.gov.au/jobs?keywords=obstetrics+gynaecology+registrar",
            "method": "playwright",
        },
        {
            "label": "RANZCOG_board",
            "portal_key": "ranzcog",
            "url": "https://jobs.ranzcog.edu.au/jobs",
            "method": "ranzcog_scraper",
        },
        {
            "label": "GV_Health_OG",
            "portal_key": "gv_health",
            "url": "https://careers.gvhealth.org.au/search/?q=obstetrics+gynaecology",
            "method": "static",
        },
        {
            "label": "AWH_OG",
            "portal_key": "awh",
            "url": "https://careers.awh.org.au/search/?q=obstetrics+gynaecology",
            "method": "static",
        },
        {
            "label": "NSW_Health_OG_broad",
            "portal_key": "pageup",
            "url": "https://jobs.health.nsw.gov.au/jobs/search?q=obstetrics+gynaecology",
            "method": "static",
        },
        {
            "label": "WAVE_OG",
            "portal_key": "wave",
            "url": "https://wave.com.au/job-listing",
            "method": "static",
        },
    ]

    for entry in searches:
        label = entry["label"]
        portal_key = entry["portal_key"]
        url = entry["url"]
        base = config.PORTAL_CONFIG.get(portal_key, {}).get("base_url", "/".join(url.split("/")[:3]))
        try:
            logger.log(f"Targeted search: {label}")
            if entry["method"] == "ranzcog_scraper":
                from scrapers.specialty_boards import scrape_ranzcog

                jobs = scrape_ranzcog()
            elif entry["method"] == "playwright":
                html = fetch_with_playwright(
                    url,
                    label=label,
                    stealth=portal_key in ("ranzcog",),
                    wait_until="domcontentloaded",
                )
                jobs = parse_job_cards(html, base, portal_key)
            else:
                html = fetch_html(url, label=label)
                jobs = parse_job_cards(html, base, portal_key)

            if entry["method"] != "ranzcog_scraper":
                if portal_key == "peninsula_health":
                    from scrapers.portal_parsers import parse_peninsula_health

                    jobs = parse_peninsula_health(html, base) or jobs
                if portal_key == "ranzcog":
                    from scrapers.portal_parsers import parse_ranzcog

                    jobs = parse_ranzcog(html, base) or jobs
                if portal_key == "wave":
                    from scrapers.portal_parsers import parse_wave

                    jobs = parse_wave(html, base) or jobs

            kept = 0
            for job in jobs:
                if not is_obgyn_job(job.title, job.specialty, job.description):
                    if not any(t in job.title.lower().replace("&amp;", "&") for t in ("obstet", "gynaec", "gynec", "o&g", "obgyn")):
                        continue
                if not job.apply_link:
                    continue
                found.append(_row_from_job(job, label))
                kept += 1
            logger.log(f"  {label}: {kept} jobs")
        except Exception as exc:  # noqa: BLE001
            logger.log(f"  {label} failed: {exc}")
    return found


def scrape_seek(logger: RunLogger) -> list[dict[str, str]]:
    """Seek Australia — O&G registrar keyword searches (static first; Playwright fallback)."""
    found: list[dict[str, str]] = []
    queries = [
        "OBGYN Registrar",
        "O&G Registrar",
        "Obstetrics Gynaecology Registrar",
        "Obstetrics and Gynaecology Registrar",
    ]
    for q in queries:
        url = f"https://www.seek.com.au/{quote_plus(q)}-jobs/in-All-Australia"
        try:
            logger.log(f"Seek: {q}")
            try:
                html = fetch_html(url, label="seek_obgyn")
            except Exception:  # noqa: BLE001
                html = fetch_with_playwright(
                    url, label="seek_obgyn", stealth=True, wait_until="domcontentloaded"
                )
            soup = soup_from_html(html)
            for a in soup.find_all("a", href=True):
                href = a.get("href", "")
                title = text(a)
                if len(title) < 8 or len(title) > 200:
                    continue
                if "/job/" not in href.lower():
                    continue
                if not is_obgyn_job(title) and not any(
                    t in title.lower() for t in ("obstet", "gynaec", "gynec", "o&g", "obgyn", "ranzcog")
                ):
                    continue
                link = absolute_url("https://www.seek.com.au", href.split("?")[0])
                found.append(_make_row(title=title, apply_link=link, portal="Seek"))
        except Exception as exc:  # noqa: BLE001
            logger.log(f"  Seek failed ({q}): {exc}")
    return found


def scrape_jora(logger: RunLogger) -> list[dict[str, str]]:
    """Jora Australia (medical listings) — O&G registrar searches (static first)."""
    found: list[dict[str, str]] = []
    queries = [
        "OBGYN Registrar",
        "O&G Registrar",
        "Obstetrics Gynaecology Registrar",
    ]
    for q in queries:
        url = f"https://au.jora.com/j?q={quote_plus(q)}&l=Australia"
        try:
            logger.log(f"Jora: {q}")
            try:
                html = fetch_html(url, label="jora_obgyn")
            except Exception:  # noqa: BLE001
                html = fetch_with_playwright(
                    url, label="jora_obgyn", stealth=True, wait_until="domcontentloaded"
                )
            soup = soup_from_html(html)
            for a in soup.find_all("a", href=True):
                href = a.get("href", "")
                title = text(a)
                if len(title) < 8 or len(title) > 200:
                    continue
                if not re.search(r"/job/|/j/", href, re.I):
                    continue
                if not is_obgyn_job(title) and not any(
                    t in title.lower() for t in ("obstet", "gynaec", "gynec", "o&g", "obgyn", "ranzcog")
                ):
                    continue
                link = absolute_url("https://au.jora.com", href.split("?")[0])
                found.append(_make_row(title=title, apply_link=link, portal="Jora"))
        except Exception as exc:  # noqa: BLE001
            logger.log(f"  Jora failed ({q}): {exc}")
    return found


def dedupe_by_url(rows: Iterable[dict[str, str]]) -> list[dict[str, str]]:
    """Keep first occurrence of each normalised Apply Link."""
    seen: set[str] = set()
    unique: list[dict[str, str]] = []
    for row in rows:
        key = normalize_url(row.get("Apply Link", ""))
        if not key or key in seen:
            continue
        seen.add(key)
        unique.append(row)
    return unique


def write_excel(rows: list[dict[str, str]], path: Path = OUTPUT_PATH) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "OBGYN_Jobs"
    ws.append(COLUMNS)
    for row in rows:
        ws.append([row.get(col, "") for col in COLUMNS])

    meta = wb.create_sheet("Run_Info")
    meta.append(["Field", "Value"])
    meta.append(["Profile", PROFILE_TAG])
    meta.append(["Generated", format_dt(now_aest())])
    meta.append(["Total jobs", len(rows)])
    meta.append(["Keywords", "; ".join(OBGYN_KEYWORDS)])
    meta.append(["Columns", ", ".join(COLUMNS)])

    for col in ws.columns:
        letter = col[0].column_letter
        width = min(max(len(str(c.value or "")) for c in col) + 2, 80)
        ws.column_dimensions[letter].width = max(width, 14)

    wb.save(path)


def main() -> int:
    logger = RunLogger()
    logger.log("=== OBGYN Job Collector started ===")
    logger.log(f"Profile: {PROFILE_TAG}")

    all_rows: list[dict[str, str]] = []

    seed = seed_from_job_tracker(logger)
    all_rows.extend(seed)

    scraped = scrape_existing_portals(logger)
    all_rows.extend(scraped)

    targeted = scrape_targeted_og_searches(logger)
    all_rows.extend(targeted)

    seek_rows = scrape_seek(logger)
    all_rows.extend(seek_rows)

    jora_rows = scrape_jora(logger)
    all_rows.extend(jora_rows)

    unique = dedupe_by_url(all_rows)
    write_excel(unique)

    logger.log(f"Raw rows: {len(all_rows)} | After URL dedupe: {len(unique)}")
    logger.log(f"Saved: {OUTPUT_PATH}")
    logger.save()

    print("=" * 56)
    print("OBGYN Job Tracker — Complete")
    print(f"Profile: {PROFILE_TAG}")
    print(f"Seed: {len(seed)} | Scraped portals: {len(scraped)} | Targeted: {len(targeted)}")
    print(f"Seek: {len(seek_rows)} | Jora: {len(jora_rows)}")
    print(f"Unique jobs (by URL): {len(unique)}")
    print(f"Excel: {OUTPUT_PATH}")
    print("=" * 56)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
