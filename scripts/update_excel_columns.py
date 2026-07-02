"""One-off: remove obsolete columns and per-profile queue sheets from Job_Tracker.xlsx."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

REPO = Path(__file__).resolve().parents[1]
EXCEL_PATH = REPO / "Job_Tracker.xlsx"

COLUMNS_TO_REMOVE = [
    "Extraction Method Used",
    "Method Reliability Note",
    "Validation Flags",
    "Notes",
]

SHEETS_TO_REMOVE = [
    "Queue_dr_gen_med_pho",
    "Queue_dr_gen_med_consultant",
    "Queue_dr_paeds_registrar",
    "Queue_dr_emergency_rmo",
    "Queue_dr_og_registrar",
]


def remove_columns_from_sheet(ws, columns: list[str]) -> int:
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    indices = sorted(
        (i + 1 for i, h in enumerate(headers) if h in columns),
        reverse=True,
    )
    for idx in indices:
        ws.delete_cols(idx)
    return len(indices)


def main() -> None:
    wb = load_workbook(EXCEL_PATH)

    removed_sheets = []
    for name in SHEETS_TO_REMOVE:
        if name in wb.sheetnames:
            wb.remove(wb[name])
            removed_sheets.append(name)

    # Remove duplicate Queue_dr_gen_med_consultant if any extra copies exist
    consultant_sheets = [s for s in wb.sheetnames if s.startswith("Queue_dr_gen_med_consultant")]
    for name in consultant_sheets:
        if name not in removed_sheets and name != "Queue_dr_gen_med_consultant":
            wb.remove(wb[name])
            removed_sheets.append(name)

    cols_removed = 0
    for sheet_name in list(wb.sheetnames):
        ws = wb[sheet_name]
        if ws.max_row < 1:
            continue
        cols_removed += remove_columns_from_sheet(ws, COLUMNS_TO_REMOVE)

    wb.save(EXCEL_PATH)
    print(f"Saved {EXCEL_PATH}")
    print(f"Removed sheets: {removed_sheets}")
    print(f"Removed {cols_removed} column instances across remaining sheets")
    print("Remaining sheets:", wb.sheetnames)
    for s in ["SmartJobs_QLD", "Apply_Queue", "All_Jobs_Australia"]:
        if s in wb.sheetnames:
            ws = wb[s]
            hdr = [ws.cell(1, c).value for c in range(1, ws.max_column + 1) if ws.cell(1, c).value]
            print(f"  {s}: {hdr}")


if __name__ == "__main__":
    main()
